from kivy.app import App
from kivy.uix.widget import Widget
from kivy.graphics import Line
from kivy.core.window import Window
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.stacklayout import StackLayout
from kivy.uix.dropdown import DropDown
from kivy.uix.stencilview import StencilView
from kivy.uix.spinner import Spinner, SpinnerOption
from kivy.uix.checkbox import CheckBox
from kivy.core.window import Window
from kivy.graphics import InstructionGroup, Color
from scipy.signal import savgol_filter
from openpyxl import load_workbook
import xlsxwriter
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from VSD_utils import *

Window.clearcolor = (1, 1, 1, 1)

red = Color(1, 0, 0)
blue = Color(0, 0, 1)
green = Color(0.149, 0.494, 0.294)
white = Color(1, 1, 1)
black = Color(0, 0, 0)
yellow = Color(1, 0.941, 0.36)
purple = Color(0.949, 0.36, 1)
orange = Color(0.827, 0.615, 0.035)

ALL_VITAL_SIGNS = ['Blood pressure', 'Blood glucose', 'Heart rate', 'Systolic', 'Diastolic']

class VitalSignInfo:
    def __init__(self, aName = ''):
        self.name = aName
        self.x_min = 0
        self.x_max = 0
        self.y_min = 0
        self.y_max = 0
        self.x_data_array = []
        self.y_data_array = []

    def isPopulated(self):
        return len(self.x_data_array) > 0 or len(self.y_data_array) > 0

class DrawInputWidget(StencilView):
    def __init__(self, **kwargs):
        super(DrawInputWidget, self).__init__(**kwargs)
        self.x_data_array = []
        self.y_data_array = []

    def on_touch_down(self, touch):
        with self.canvas:
            touch.ud["line"] = Line(points=(touch.x, touch.y), width = 2)

    def on_touch_move(self, touch):
        # check if the point is on the drawing area
        if self.collide_point(touch.x, touch.y):
            touch.ud["line"].points += (touch.x, touch.y)
            self.x_data_array.append(touch.x)
            self.y_data_array.append(touch.y)

    def on_touch_up(self, touch):
        print("RELEASED!",touch)


# We want to map the interval (a, b) one to one to (c, d)
# Returning the coefficients in f(x) = mx + n
def mapIntervals(a, b, c, d):
    m  = (c - d) / (a - b)
    n = c - m * a
    return m, n


class ParentLayout(StackLayout):
    def __init__(self, **kwargs):
        super(ParentLayout, self).__init__(**kwargs)
        self.output_file_name = 'Output file name'
        self.vital_sign = ''
        self.x_min = 0
        self.x_max = 0
        self.y_min = 0
        self.y_max = 0
        self.data = dict()
        for item in ALL_VITAL_SIGNS:
            newInfo = VitalSignInfo(item)
            self.data[newInfo.name] = newInfo

        self.ids.save_img_chkbox.bind(active = self.on_save_img_chkbox_active)
        self.save_image_flag = True


    def on_save_img_chkbox_active(self, checkbox, active):
        self.save_image_flag = active

    # Scale data linearly in each x and y component
    def scale_data(self):
        for item in ALL_VITAL_SIGNS:
            current_data = self.data[item]
            if current_data.isPopulated():
                x_scaled_array = []
                y_scaled_array = []

                m_x, n_x = mapIntervals(min(current_data.x_data_array), max(current_data.x_data_array), current_data.x_min, current_data.x_max)
                m_y, n_y = mapIntervals(min(current_data.y_data_array), max(current_data.y_data_array), current_data.y_min, current_data.y_max)
                for x in current_data.x_data_array:
                    scaledX = m_x * x + n_x
                    x_scaled_array.append(scaledX)
                for y in current_data.y_data_array:
                    scaledY = m_y * y + n_y
                    y_scaled_array.append(scaledY)

                current_data.x_data_array = x_scaled_array
                current_data.y_data_array = y_scaled_array

    # Smooth the data for better chart drawing
    def smooth_data(self):
        dof = 6

        for item in ALL_VITAL_SIGNS:
            current_data = self.data[item]
            if current_data.isPopulated():
                current_y_array = current_data.y_data_array
                window_size = min(51, len(current_y_array) )
                # window_size must be odd
                if (window_size % 2 == 0):
                    window_size = window_size - 1
                current_y_array = savgol_filter(current_y_array, window_size, dof)
                # Smoothing data may make value exceeds thresholds, so we need to adjust this
                current_y_array[current_y_array > current_data.y_max] = current_data.y_max
                current_y_array[current_y_array < current_data.y_min] = current_data.y_min

    def print_data_to_output_file(self):
        fileName = self.output_file_name + '.xlsx'

        # Create a new excel file
        workbook = xlsxwriter.Workbook(fileName)
        workbook.close()

        book = load_workbook(fileName)
        with pd.ExcelWriter(fileName, engine='openpyxl') as writer:
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            for item in ALL_VITAL_SIGNS:
                current_data = self.data[item]
                if current_data.isPopulated():
                    pd.DataFrame(current_data.y_data_array).to_excel(writer, sheet_name = current_data.name)
            writer.save()

    # this functions will validate if all the fields are populated / valid
    def validate_before_applying(self):
        err_msg = ''

        #check if output_file_name is already entered
        if self.output_file_name == '':
            err_msg = 'Output file name is not entered !!!'
            show_message_box(err_msg, 'Error')
            return False

        num_of_chosen_signs = 0
        for item in ALL_VITAL_SIGNS:
            current_data = self.data[item]
            if current_data.isPopulated():
                if (current_data.x_min < 0) or (current_data.x_max < 0) or (current_data.y_min < 0) or (current_data.y_max < 0):
                    show_message_box('Ranges cannot be negative !!!', 'Error')
                    return False
                elif (current_data.x_min >= current_data.x_max):
                   err_msg = 'x_min >= x_max for ' + current_data.name
                   show_message_box(err_msg, 'Error')
                   return False
                elif (current_data.y_min >= current_data.y_max):
                   err_msg = 'x_min >= x_max for ' + current_data.name
                   show_message_box(err_msg, 'Error')
                   return False
                num_of_chosen_signs += 1

        if (num_of_chosen_signs == 0) and (self.vital_sign == ''):
            show_message_box('None of the vital signs is chosen !!!', 'Error')
            return False

        return True

    def apply_data(self, obj):
        # make sure to add the last vital_sign choice
        current_data = self.data[self.vital_sign]
        if not current_data.isPopulated():
            newInfo = VitalSignInfo(self.vital_sign)
            newInfo.x_data_array = self.ids.draw_area.x_data_array
            newInfo.y_data_array = self.ids.draw_area.y_data_array
            newInfo.x_max = float(self.ids.x_max.text)
            newInfo.x_min = float(self.ids.x_min.text)
            newInfo.y_min = float(self.ids.y_min.text)
            newInfo.y_max = float(self.ids.y_max.text)
            self.data[self.vital_sign] = newInfo

        # validate the options first
        if self.validate_before_applying():
            # Scale data
            self.scale_data()

            # Smoothing the data
            self.smooth_data()

            # Print data to the output file
            self.print_data_to_output_file()
            print('Data applied to ' + self.output_file_name)

            # Save the image
            if self.save_image_flag:
                self.save_image()

    def save_image(self):
        num_of_chosen_signs = 0
        for item in ALL_VITAL_SIGNS:
            current_data = self.data[item]
            if current_data.isPopulated():
                num_of_chosen_signs += 1

        fig, axs = plt.subplots(num_of_chosen_signs)
        cntr = 0
        for item in ALL_VITAL_SIGNS:
            current_data = self.data[item]
            if current_data.isPopulated():
               axs[cntr].plot(current_data.x_data_array, current_data.y_data_array)
               axs[cntr].set_title(current_data.name)
               cntr += 1

        fig.savefig(self.output_file_name + '.png')
        plt.show(fig)

    def output_file_text_changed_handler(self, instance, text):
        self.output_file_name = text
        return text

    def clear_btn_pressed_handler(self, obj):
        print('Clearing Screen!')
        self.ids.draw_area.canvas.clear()
        self.obj = InstructionGroup()
        self.obj.add(self.vital_sign_to_color(self.vital_sign))
        self.ids.draw_area.canvas.add(self.obj)

        # reset the data
        for item in ALL_VITAL_SIGNS:
            newInfo = VitalSignInfo(item)
            self.data[newInfo.name] = newInfo

        # Reset range fields
        self.ids.x_min.text = '0'
        self.ids.x_max.text = '0'
        self.ids.y_min.text = '0'
        self.ids.y_max.text = '0'

        # Reset cached data
        self.ids.draw_area.x_data_array = []
        self.ids.draw_area.y_data_array = []

    def axis_range_changed(self, instance, text, id):
        num = float(text)

        if (id == 'x_max'):
            self.data[self.vital_sign].x_max = num
            print('x_max changed to ' + str(self.data[self.vital_sign].x_max))
        elif (id == 'x_min'):
            self.data[self.vital_sign].x_min = num
            print('x_min changed to ' + str(self.data[self.vital_sign].x_min))
        elif (id == 'y_max'):
            self.data[self.vital_sign].y_max = num
            print('y_max changed to ' + str(self.data[self.vital_sign].y_max))
        elif (id == 'y_min'):
            self.data[self.vital_sign].y_min = num
            print('y_min changed to ' + str(self.data[self.vital_sign].y_min))

    def on_vital_sign_spinner_select(self, text):
        if (text != 'Choose one'):
            if self.vital_sign != '':
                self.data[self.vital_sign].x_data_array = self.ids.draw_area.x_data_array
                self.data[self.vital_sign].y_data_array = self.ids.draw_area.y_data_array

            self.vital_sign = text
            self.obj = InstructionGroup()
            self.obj.add(self.vital_sign_to_color(text))
            self.ids.draw_area.canvas.add(self.obj)

            # Update the range fields
            self.ids.x_min.text = str(self.data[self.vital_sign].x_min)
            self.ids.x_max.text = str(self.data[self.vital_sign].x_max)
            self.ids.y_min.text = str(self.data[self.vital_sign].y_min)
            self.ids.y_max.text = str(self.data[self.vital_sign].y_max)

        print(self.vital_sign + ' is selected')

        # Clear current data
        self.ids.draw_area.x_data_array = []
        self.ids.draw_area.y_data_array = []

    def vital_sign_to_color(self, text):
        if (text == 'Blood pressure'):
            return blue
        elif (text == 'Blood glucose'):
            return orange
        elif (text == 'Heart rate'):
            return red
        elif (text == 'Systolic'):
            return green
        elif (text == 'Diastolic'):
            return purple
        else:
            return Window.clearcolor

class VitalSignDrawer(App):
    def build(self):
        return ParentLayout()

if __name__ == '__main__':
    VitalSignDrawer().run()